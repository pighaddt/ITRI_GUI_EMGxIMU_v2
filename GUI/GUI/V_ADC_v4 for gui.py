from numpy import *
import numpy as np
import serial
import openpyxl
import time



# Create object serial port
file = 722005
portName = "COM3"
baudRate = 115200
ser = serial.Serial(portName, int(baudRate), timeout=1, parity=serial.PARITY_NONE, stopbits=1)

# samplePoint = 2000    # after how many points -> starting cal C
# thrd = 30        # how many periods would be cal
# tt = 15          # measure 10 sec

samplePoint = 1000    # after how many points -> starting cal C
thrd = 10        # how many periods would be cal
tt = 3           # measure 10 sec

# Parameter setting
R = 1 * 10**6             # Ohm
pwmPeriod = 2 * 10**(-3) #sec
sampleHZ = 6 * 10**3     # Hz

sampleRate = 1/sampleHZ   # sec
sampleRate_p = int(sampleRate // 10**(-6))  # micro-sec
point = pwmPeriod//sampleRate
tauPercent = 0.632



# Create a form in Excel
workbook = openpyxl.Workbook()
sheet = workbook.create_sheet('V_ADC test1')


#####
def update(sample):
    """
    Each time this function is called, the data display is updated
    """
    Data = []
    print(time.ctime())
    for i in range(sample):
        value = str(ser.readline(), 'utf-8')
        if value != '\n':
            # print(value)
            Data.append(int(value)/1000)
    print(Data)
    return Data


#####
def calcuC(Value):
    """
    The function of calculating capacitance.
    """
    filterValue = Value

    # Find the minimum/maximun value of the curve by calculating the larger/smaller difference.
    diff, MaxIndx, MaxIndxValue, MinIndx, MinIndxValue = [], [], [], [], []
    for i in range(len(filterValue) - 1):
        diff.append(filterValue[i + 1] - filterValue[i])
    for i in range(thrd):
        maxIndx = np.argmax(diff)
        minIndx = np.argmin(diff)
        if maxIndx + int(point // 2) < len(filterValue):
            if filterValue[maxIndx + int(point // 2)] > filterValue[maxIndx + int((point // 2) // 2)]:
                if filterValue[maxIndx + 1] > filterValue[maxIndx] and filterValue[maxIndx - 1] > filterValue[maxIndx]:
                    MaxIndx.append(maxIndx)
                    MaxIndxValue.append(filterValue[maxIndx])
        if minIndx + int(point // 2) < len(filterValue):
            if filterValue[minIndx + int(point // 2)] < filterValue[minIndx + int((point // 2) // 2)]:
                if filterValue[minIndx + 1] < filterValue[minIndx] and filterValue[minIndx - 1] < filterValue[minIndx]:
                    MinIndx.append(minIndx)
                    MinIndxValue.append(filterValue[minIndx])
        diff[maxIndx] = 0
        diff[minIndx] = 0

    # Delete the error items of maximum/minimum value
    for i in range(len(MaxIndx)):
        for j in range(1, int(point // 2) - 1):
            if filterValue[MaxIndx[i] + j] < filterValue[(MaxIndx[i] + 1)]:
                MaxIndx[i] = 0
                MaxIndxValue[i] = 0
    if MaxIndx.count(0):
        if 0 in MaxIndx:
            MaxIndx.remove(0)
            # MaxIndxValue.remove(0)

    for i in range(len(MinIndx)):
        for j in range(1, int(point // 2) - 1):
            if filterValue[MinIndx[i] + j] > filterValue[(MinIndx[i] + 1)]:
                MinIndx[i] = 0
                MinIndxValue[i] = 0
    if MinIndx.count(0):
        if 0 in MinIndx:
            MinIndx.remove(0)
            # MinIndxValue.remove(0)

    ###
    # case 1: rise
    C_rise = []
    for p in range(len(MaxIndx)):
        V_tau = (filterValue[MaxIndx[p] + int(point // 2)] - filterValue[MaxIndx[p]]) * tauPercent + filterValue[
            MaxIndx[p]]
        for i in range(int(point // 2) + 1):
            if V_tau < filterValue[MaxIndx[p] + i]:
                tauIndex = i - 1
                break
        t_tau = float((V_tau - filterValue[MaxIndx[p] + tauIndex]) / (
                    filterValue[MaxIndx[p] + tauIndex + 1] - filterValue[
                MaxIndx[p] + tauIndex])) * sampleRate + tauIndex * sampleRate
        C_rise.append(abs(t_tau) / R)


    # case 2: down
    C_down = []
    for p in range(len(MinIndx)):
        V_tau = (filterValue[MinIndx[p]] - (filterValue[MinIndx[p]] - filterValue[MinIndx[p] + int(point // 2)]) * tauPercent)
        for i in range(int(point // 2) + 1):
            if V_tau > filterValue[MinIndx[p] + i]:
                tauIndex = i - 1
                break
        t_tau = float((filterValue[MinIndx[p] + tauIndex] - V_tau) / (filterValue[MinIndx[p] + tauIndex] - filterValue[MinIndx[p] + tauIndex + 1])) * sampleRate + tauIndex * sampleRate
        C_down.append(abs(t_tau) / R)
        C_rise.append(abs(t_tau) / R)

    dic = {'time': time.ctime(), 'Length of C': len(C_rise), 'Average capacitance value(F)': np.average(C_rise)}
    print(dic)
    print()
    return len(C_rise), C_rise, dic


f = open('V_ADC Data_{}.txt'.format(file), 'w')
###### Start ######
for t in range(tt):
    # load signal
    print("##### No.{} #####".format(t+1))
    Data = update(samplePoint)

    # calculate C
    C_len, CalC, message = calcuC(Data)

    # Save the Text file
    imform = [str(message), " , All capacitance value(F): ", str(CalC), "\n"]
    f.writelines(imform)

    # write the data into the Excel file
    for i in range(len(Data)):
        sheet.cell(row=int(i+1), column=t+1, value=Data[i])


# Save the Excel file
workbook.save('V_ADC Data_{}.xlsx'.format(file))
f.close()